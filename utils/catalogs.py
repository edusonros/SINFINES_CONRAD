# utils/catalogs.py
import os
from typing import Dict, List, Any, Optional

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_XLSX = os.path.join(BASE_DIR, "data", "Lista de Planos.xlsx")


def _fmt_mm(x) -> str:
    """Formatea mm a string estilo ES: '60,3' y sin ',0'."""
    if x is None:
        return ""
    try:
        f = float(x)
    except Exception:
        return str(x).strip()

    if abs(f - round(f)) < 1e-9:
        return str(int(round(f)))
    s = f"{f:.1f}".replace(".", ",")
    return s


def _col_values(ws, col: int, start_row: int = 2) -> List[Any]:
    out = []
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None or str(v).strip() == "":
            continue
        out.append(v)
    return out


def load_catalogs(xlsx_path: str = DEFAULT_XLSX) -> Dict[str, Any]:
    """
    Devuelve un dict con listas para combos:
      - materials: [str]
      - eje_od: [str] (Ø exterior tubo eje)
      - espesores_by_od: { "60,3": ["2", "3", "4"...], ... }
      - diam_espira: [str]
      - pasos: [str]
      - espesores_chapa: [str]
      - tipo_disposicion: [str] (COAXIAL/ORTOGONAL/PENDULAR/...)
      - posicion_motor: [str] (B3/B6/...)
      - giro: [str] (DERECHAS/IZQUIERDAS)
      - rodamientos: [ {ref, d, D, B}, ... ]
      - eje_dim: [str] (diámetros d disponibles en rodamientos; útil para mangones/árbol)
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(
            f"No existe el Excel de catálogos: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ===== materiales =====
    materials: List[str] = []
    if "LISTA_MATERIALES" in wb.sheetnames:
        ws = wb["LISTA_MATERIALES"]
        materials = [str(v).strip() for v in _col_values(ws, 1, 2)]
        materials = [m for m in materials if m]

    # ===== eje tubo: Ø exterior + espesor =====
    eje_od_set = set()
    espesores_by_od: Dict[str, List[str]] = {}
    if "LISTA_EJE_TUBO" in wb.sheetnames:
        ws = wb["LISTA_EJE_TUBO"]
        for r in range(2, ws.max_row + 1):
            od = ws.cell(r, 1).value
            thk = ws.cell(r, 2).value
            if od is None or thk is None:
                continue
            try:
                od_f = float(od)
                thk_f = float(thk)
            except Exception:
                continue
            od_s = _fmt_mm(od_f)
            thk_s = _fmt_mm(thk_f)
            eje_od_set.add(od_s)
            espesores_by_od.setdefault(od_s, [])
            if thk_s not in espesores_by_od[od_s]:
                espesores_by_od[od_s].append(thk_s)

        # ordenar espesores numéricamente
        for k in espesores_by_od:
            espesores_by_od[k] = sorted(
                espesores_by_od[k],
                key=lambda s: float(s.replace(",", ".")) if s else 0.0
            )

    eje_od = sorted(
        list(eje_od_set),
        key=lambda s: float(s.replace(",", ".")) if s else 0.0
    )

    # ===== diametros, pasos, chapa, disposicion, posicion, giro =====
    diam_espira: List[str] = []
    pasos: List[str] = []
    espesores_chapa: List[str] = []
    tipo_disposicion: List[str] = []
    posicion_motor: List[str] = []
    giro: List[str] = []

    if "DIAMETROS Y PASOS" in wb.sheetnames:
        ws = wb["DIAMETROS Y PASOS"]

        # columnas según tu Excel:
        # A: diam espira, B: pasos, C: espesor chapa, E: tipo disposición, F: posición, G: giro
        diam_espira = sorted(
            {_fmt_mm(v) for v in _col_values(ws, 1, 2)
             if v is not None and str(v).strip() != ""},
            key=lambda s: float(s.replace(",", ".")) if s else 0.0
        )

        pasos = sorted(
            {_fmt_mm(v) for v in _col_values(ws, 2, 2)
             if v is not None and str(v).strip() != ""},
            key=lambda s: float(s.replace(",", ".")) if s else 0.0
        )

        espesores_chapa = sorted(
            {_fmt_mm(v) for v in _col_values(ws, 3, 2)
             if v is not None and str(v).strip() != ""},
            key=lambda s: float(s.replace(",", ".")) if s else 0.0
        )

        tipo_disposicion = sorted(
            {str(v).strip() for v in _col_values(ws, 5, 2)
             if v is not None and str(v).strip() != ""}
        )

        posicion_motor = sorted(
            {str(v).strip() for v in _col_values(ws, 6, 2)
             if v is not None and str(v).strip() != ""}
        )

        giro = sorted(
            {str(v).strip() for v in _col_values(ws, 7, 2)
             if v is not None and str(v).strip() != ""}
        )

    # ===== rodamientos =====
    rodamientos: List[Dict[str, Any]] = []
    eje_dim_set = set()
    if "CAT_RODAMIENTOS" in wb.sheetnames:
        ws = wb["CAT_RODAMIENTOS"]
        # Estructura típica:
        # A: REF, B: ?, C: d, D: D, E: B  (en tu excel d está en col 3)
        for r in range(2, ws.max_row + 1):
            ref = ws.cell(r, 1).value
            d = ws.cell(r, 3).value
            D = ws.cell(r, 4).value
            B = ws.cell(r, 5).value
            if not ref or d is None:
                continue
            try:
                d_f = float(d)
                D_f = float(D) if D is not None else None
                B_f = float(B) if B is not None else None
            except Exception:
                continue

            item = {
                "ref": str(ref).strip(),
                "d": d_f,
                "D": D_f,
                "B": B_f,
            }
            rodamientos.append(item)
            eje_dim_set.add(_fmt_mm(d_f))

        rodamientos.sort(key=lambda x: (x["d"], x["ref"]))

    eje_dim = sorted(
        list(eje_dim_set),
        key=lambda s: float(s.replace(",", ".")) if s else 0.0
    )

    return {
        "materials": materials,
        "eje_od": eje_od,
        "espesores_by_od": espesores_by_od,
        "diam_espira": diam_espira,
        "pasos": pasos,
        "espesores_chapa": espesores_chapa,
        "tipo_disposicion": tipo_disposicion,
        "posicion_motor": posicion_motor,
        "giro": giro,
        "rodamientos": rodamientos,
        "eje_dim": eje_dim,
    }


def filter_rodamientos_por_d(rodamientos: List[Dict[str, Any]], d_mm_text: str) -> List[Dict[str, Any]]:
    """Filtra rodamientos por diámetro interior exacto (d)."""
    if not d_mm_text:
        return rodamientos
    try:
        d = float(str(d_mm_text).replace(",", "."))
    except Exception:
        return rodamientos
    return [r for r in rodamientos if abs(r["d"] - d) < 1e-6]


def filter_rodamientos_por_eje(rodamientos: list, eje_od_mm_text: str) -> list:
    """
    Devuelve una lista de referencias (str) filtradas según el Ø del tubo eje.
    - Si tus rodamientos en el catálogo no tienen campo 'd' (diámetro interior), devuelve todas las refs.
    - Si tienen 'd', filtra por d < Ø_eje (regla simple; luego la afinamos).
    """
    if not rodamientos:
        return []

    eje_od_mm_text = (eje_od_mm_text or "").strip().replace(",", ".")
    try:
        eje_od = float(eje_od_mm_text)
    except Exception:
        # no puedo comparar -> devuelvo todas las refs
        return [r.get("ref") for r in rodamientos if r.get("ref")]

    refs = []
    for r in rodamientos:
        ref = r.get("ref")
        if not ref:
            continue

        d = r.get("d", None)
        if d is None:
            # no hay dato para filtrar
            refs.append(ref)
            continue

        try:
            d_val = float(str(d).replace(",", "."))
        except Exception:
            refs.append(ref)
            continue

        # filtro simple (luego lo ajustamos a tu lógica real)
        if d_val < eje_od:
            refs.append(ref)

    # quitar duplicados manteniendo orden
    out = []
    seen = set()
    for x in refs:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out
